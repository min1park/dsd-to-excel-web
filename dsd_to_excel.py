#!/usr/bin/env python3
"""
DSD → Excel 변환기
DART 전자공시 감사보고서(.dsd) 파일을 Excel(.xlsx) 워크북으로 변환합니다.
"""

import sys
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# 상수
# ============================================================================
DEFAULT_FONT_NAME = '맑은 고딕'
DEFAULT_FONT_SIZE = 8
THIN_BORDER_SIDE = Side(style='thin')
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE
)
NUMBER_FORMAT = '#,##0_);\\(#,##0\\);\\-_)'
FS_DIVIDER = 1000
DATA_COL_START = 4  # D열


# ============================================================================
# 1단계: DSD 파싱
# ============================================================================

def parse_dsd(filepath):
    """DSD 파일(ZIP)에서 contents.xml과 meta.xml을 파싱하여 반환"""
    with zipfile.ZipFile(filepath, 'r') as zf:
        meta_xml_str = zf.read('meta.xml').decode('utf-8')
        meta_root = ET.fromstring(meta_xml_str)

        contents_xml_str = zf.read('contents.xml').decode('utf-8')
        contents_xml_str = contents_xml_str.replace('&cr;', '\n')
        contents_root = ET.fromstring(contents_xml_str)

    return meta_root, contents_root


# ============================================================================
# 2단계: XML 유틸리티
# ============================================================================

def get_text(elem):
    """요소와 모든 하위 요소의 텍스트를 합쳐서 반환.
    리터럴 '&cr;' 문자열도 줄바꿈으로 치환."""
    if elem is None:
        return ''
    parts = []
    if elem.text:
        parts.append(elem.text)
    for child in elem:
        parts.append(get_text(child))
        if child.tail:
            parts.append(child.tail)
    result = ''.join(parts)
    # XML 레벨에서 치환되지 않은 리터럴 &cr; 처리
    result = result.replace('&cr;', '\n')
    return result


def parse_usermark(usermark):
    """USERMARK 문자열에서 폰트 정보 추출"""
    if not usermark:
        return {}
    info = {}
    tokens = usermark.strip().split()
    for token in tokens:
        if token == 'B':
            info['bold'] = True
        elif token == '!B':
            info['bold'] = False
        elif token.startswith('F-') and token != 'F-GL':
            size_str = token[2:].replace('BT', '')
            try:
                info['size'] = int(size_str)
            except ValueError:
                pass
        elif token.startswith('P') and token[1:].isdigit():
            try:
                info['size'] = int(token[1:])
            except ValueError:
                pass
        elif token.startswith('0X') or token.startswith('0x'):
            color = token[2:]
            if len(color) == 6:
                info['color'] = color
    return info


def is_valid_thousands_format(s):
    """천단위 구분 쉼표가 올바른 형식인지 검증.
    올바른 형식: 1,234 / 12,345,678 / 1,234,567
    잘못된 형식: 6,25,29 / 1,23 / ,123"""
    if ',' not in s:
        return True
    # 음수 부호 제거
    s = s.lstrip('-')
    parts = s.split(',')
    if not parts[0] or not parts[0].isdigit():
        return False
    # 첫 번째 부분: 1~3자리
    if len(parts[0]) > 3 or len(parts[0]) == 0:
        return False
    # 나머지 부분: 정확히 3자리
    for part in parts[1:]:
        if len(part) != 3 or not part.isdigit():
            return False
    return True


def try_parse_number(text):
    """텍스트에서 숫자를 파싱 시도. (값, True) 또는 (텍스트, False)"""
    if not text:
        return text, False
    cleaned = text.strip()
    if not cleaned or cleaned == '\u3000':
        return cleaned, False
    # 대시(-)는 숫자 0으로 처리 (재무제표에서 값 없음 표시)
    if cleaned == '-':
        return 0, True

    # 괄호 음수: (1,234,567)
    neg_match = re.match(r'^\(([0-9,]+)\)$', cleaned)
    if neg_match:
        inner = neg_match.group(1)
        if not is_valid_thousands_format(inner):
            return cleaned, False
        num_str = inner.replace(',', '')
        try:
            return -int(num_str), True
        except ValueError:
            return cleaned, False

    # 일반 정수: 1,234,567
    num_match = re.match(r'^-?[0-9,]+$', cleaned)
    if num_match:
        if not is_valid_thousands_format(cleaned):
            return cleaned, False
        num_str = cleaned.replace(',', '')
        try:
            return int(num_str), True
        except ValueError:
            return cleaned, False

    # 소수: 1,234.56
    dec_match = re.match(r'^-?([0-9,]+)\.[0-9]+$', cleaned)
    if dec_match:
        int_part = dec_match.group(1)
        if not is_valid_thousands_format(int_part.lstrip('-')):
            return cleaned, False
        num_str = cleaned.replace(',', '')
        try:
            return float(num_str), True
        except ValueError:
            return cleaned, False

    return cleaned, False


def px_to_excel_width(px):
    """픽셀 너비를 Excel 열 너비 단위로 변환"""
    return max(px / 7.0, 2.0)


# ============================================================================
# 3단계: 테이블 파싱
# ============================================================================

def parse_table(table_elem):
    """TABLE XML 요소를 파싱하여 구조화된 딕셔너리 반환"""
    result = {
        'border': table_elem.get('BORDER', '0') == '1',
        'width': int(table_elem.get('WIDTH', '600')),
        'aclass': table_elem.get('ACLASS', 'NORMAL'),
        'col_widths': [],
        'thead_rows': [],
        'tbody_rows': [],
    }

    colgroup = table_elem.find('COLGROUP')
    if colgroup is not None:
        for col in colgroup.findall('COL'):
            result['col_widths'].append(int(col.get('WIDTH', '100')))

    thead = table_elem.find('THEAD')
    if thead is not None:
        for tr in thead.findall('TR'):
            result['thead_rows'].append(parse_tr(tr, is_header=True))

    tbody = table_elem.find('TBODY')
    if tbody is not None:
        for tr in tbody.findall('TR'):
            result['tbody_rows'].append(parse_tr(tr, is_header=False))

    return result


def parse_tr(tr_elem, is_header=False):
    """TR 요소를 파싱하여 셀 목록 반환"""
    cells = []
    for cell_elem in tr_elem:
        tag = cell_elem.tag
        if tag not in ('TH', 'TD', 'TE', 'TU'):
            continue
        cell = {
            'tag': tag,
            'text': get_text(cell_elem).strip(),
            'colspan': int(cell_elem.get('COLSPAN', '1')),
            'rowspan': int(cell_elem.get('ROWSPAN', '1')),
            'align': (cell_elem.get('ALIGN') or '').upper(),
            'valign': (cell_elem.get('VALIGN') or '').upper(),
            'usermark': cell_elem.get('USERMARK', ''),
            'width': int(cell_elem.get('WIDTH', '0')),
            'is_header': tag == 'TH' or is_header,
        }
        cells.append(cell)
    return cells


# ============================================================================
# 4단계: 문서 구조 분석
# ============================================================================

def analyze_document(root):
    """XML 문서 구조를 분석하여 섹션별 콘텐츠 매핑"""
    doc = {
        'header': {},
        'cover': None,
        'toc': None,
        'opinion_section': None,
        'fs_section': None,
        'notes_section': None,
        'conduct_section': None,
    }

    dh = root.find('DOCUMENT-HEADER')
    if dh is not None:
        dn = dh.find('DOCUMENT-NAME')
        cn = dh.find('COMPANY-NAME')
        doc['header']['doc_name'] = get_text(dn) if dn is not None else ''
        doc['header']['company_cik'] = cn.get('AREGCIK', '') if cn is not None else ''

    body = root.find('BODY')
    if body is None:
        print("ERROR: BODY 요소를 찾을 수 없습니다.")
        sys.exit(1)

    for child in body:
        tag = child.tag

        if tag == 'COVER':
            doc['cover'] = child

        elif tag == 'TOC':
            doc['toc'] = child

        elif tag == 'INSERTION':
            if child.get('AFREQUENCY') == '1':
                lib = child.find('LIBRARY')
                if lib is not None:
                    sec1 = lib.find('SECTION-1')
                    if sec1 is not None:
                        title_elem = sec1.find('TITLE')
                        title_text = get_text(title_elem) if title_elem is not None else ''
                        if '감사' in title_text and '보고서' in title_text:
                            doc['opinion_section'] = sec1

        elif tag == 'SECTION-1':
            title_elem = child.find('TITLE')
            title_text = get_text(title_elem) if title_elem is not None else ''

            if '재' in title_text and '무' in title_text and '표' in title_text:
                doc['fs_section'] = child
                for sec2 in child.findall('SECTION-2'):
                    sec2_title = sec2.find('TITLE')
                    sec2_text = get_text(sec2_title) if sec2_title is not None else ''
                    if '주석' in sec2_text:
                        doc['notes_section'] = sec2

            elif '외부감사' in title_text:
                doc['conduct_section'] = child

    return doc


def extract_cover_info(cover_elem):
    """COVER 요소에서 표지 정보 추출"""
    info = {
        'company_name': '',
        'report_subtitle': '',
        'report_title': '',
        'period_line1': '',
        'period_from': '',
        'period_to': '',
        'auditor_name': '',
    }
    if cover_elem is None:
        return info

    tables = cover_elem.findall('TABLE')
    cover_title = cover_elem.find('COVER-TITLE')
    table_groups = cover_elem.findall('TABLE-GROUP')

    table_texts = []
    for t in tables:
        txt = get_text(t).strip()
        if txt:
            table_texts.append(txt)

    if len(table_texts) >= 1:
        info['company_name'] = table_texts[0]
    if len(table_texts) >= 2:
        info['report_subtitle'] = table_texts[1]
    if len(table_texts) >= 3:
        info['auditor_name'] = table_texts[2]

    if cover_title is not None:
        info['report_title'] = get_text(cover_title).strip()

    for tg in table_groups:
        for table in tg.findall('TABLE'):
            tbody = table.find('TBODY')
            if tbody is None:
                continue
            for tr in tbody.findall('TR'):
                for cell in tr:
                    if cell.tag == 'TU':
                        aunit = cell.get('AUNIT', '')
                        text = get_text(cell).strip()
                        if aunit == 'PERIODFROM':
                            info['period_from'] = text
                        elif aunit == 'PERIODTO':
                            info['period_to'] = text
                    elif cell.tag == 'TD':
                        text = get_text(cell).strip()
                        if '기' in text and ('당' in text or '전' in text):
                            info['period_line1'] = text

    return info


def extract_toc_items(toc_elem):
    """TOC에서 목차 항목 추출"""
    items = []
    if toc_elem is None:
        return items

    title_elem = toc_elem.find('TITLE')
    if title_elem is not None:
        items.append(('title', get_text(title_elem).strip()))

    for table in toc_elem.findall('TABLE'):
        tbody = table.find('TBODY')
        if tbody is None:
            continue
        for tr in tbody.findall('TR'):
            cells = tr.findall('TD')
            if len(cells) >= 2:
                left = get_text(cells[0]).strip()
                right = get_text(cells[1]).strip()
                if left:
                    items.append(('entry', left, right))
            elif len(cells) == 1:
                text = get_text(cells[0]).strip()
                if text:
                    items.append(('entry', text, ''))
    return items


def split_financial_statements(fs_section):
    """재무제표 SECTION-1에서 개별 재무제표 분리.

    구조: 각 재무제표는 3개 TABLE + PGBRK 패턴:
      TABLE(BORDER=0, 5행, 제목) + TABLE(BORDER=1, 데이터) + TABLE(BORDER=0, 1행, 각주) + PGBRK
    """
    statements = {
        'fs_header': [],    # 재무제표 표지 (SECTION-1 첫 부분)
        'bs': [],           # 재무상태표 요소들
        'is': [],           # 포괄손익계산서 요소들
        'ce': [],           # 자본변동표 요소들
        'cf': [],           # 현금흐름표 요소들
    }

    if fs_section is None:
        return statements

    # SECTION-2(주석) 제외한 직접 자식 수집
    children = []
    for child in fs_section:
        if child.tag == 'SECTION-2':
            continue
        children.append(child)

    # 재무제표 데이터 TABLE(BORDER=1) 위치 찾기
    border1_indices = []
    for i, child in enumerate(children):
        if child.tag == 'TABLE' and child.get('BORDER') == '1':
            border1_indices.append(i)

    # 각 BORDER=1 테이블의 THEAD로 재무제표 유형 판별
    fs_blocks = []  # (유형, 시작인덱스, 끝인덱스)
    for bi in border1_indices:
        data_table = children[bi]
        thead = data_table.find('THEAD')
        if thead is None:
            continue

        # 제목 TABLE = 바로 앞 TABLE (BORDER=0)
        title_idx = bi - 1
        while title_idx >= 0 and children[title_idx].tag in ('P', 'WARNING', 'INSERTION'):
            title_idx -= 1

        # 제목 텍스트로 유형 판별
        title_text = ''
        if title_idx >= 0 and children[title_idx].tag == 'TABLE':
            title_text = get_text(children[title_idx]).strip()

        # 각주 TABLE = 바로 뒤
        footer_idx = bi + 1

        # 유형 판별
        fs_type = None
        if '재무상태표' in title_text.replace(' ', ''):
            fs_type = 'bs'
        elif '손익계산서' in title_text.replace(' ', ''):
            fs_type = 'is'
        elif '자본변동표' in title_text.replace(' ', ''):
            fs_type = 'ce'
        elif '현금흐름표' in title_text.replace(' ', ''):
            fs_type = 'cf'

        if fs_type:
            # 제목 TABLE ~ 각주 TABLE까지
            start = title_idx
            end = footer_idx
            # 각주 TABLE 확인
            if end < len(children) and children[end].tag == 'TABLE':
                footer_text = get_text(children[end]).strip()
                if '주석' in footer_text or '별첨' in footer_text:
                    end = end  # 각주 포함
                else:
                    end = bi  # 데이터까지만
            fs_blocks.append((fs_type, start, end))

    # 블록 인덱스 셋
    block_indices = set()
    for fs_type, start, end in fs_blocks:
        for i in range(start, end + 1):
            block_indices.add(i)

    # fs_header: 첫 번째 BORDER=1 전까지 (INSERTION/WARNING/PGBRK 제외)
    first_data = border1_indices[0] if border1_indices else len(children)
    # 첫 번째 재무제표 제목 TABLE 전까지가 헤더
    first_title = fs_blocks[0][1] if fs_blocks else first_data
    for i in range(0, first_title):
        child = children[i]
        if child.tag in ('INSERTION', 'WARNING', 'PGBRK', 'TITLE'):
            continue
        statements['fs_header'].append(child)

    # 각 재무제표 블록 할당
    for fs_type, start, end in fs_blocks:
        for i in range(start, end + 1):
            if i < len(children):
                statements[fs_type].append(children[i])

    return statements


def split_notes(notes_section, max_sheet_num=33):
    """주석 SECTION-2를 주석 번호(N. 제목) 기준으로 분리.

    원본 XLSM에서 각 주석 시트는 "N. 제목" P요소로 시작됨.
    max_sheet_num 이후의 주석은 마지막 시트에 합침.
    반환: [(시트번호, [요소들]), ...]
    """
    if notes_section is None:
        return []

    # 모든 자식 수집 (TITLE 제외, PGBRK 포함)
    children = []
    for child in notes_section:
        if child.tag == 'TITLE':
            continue
        children.append(child)

    # 주석 번호 시작점 찾기: "N. " 패턴으로 시작하는 P 요소
    note_starts = []  # (인덱스, 번호)
    note_number_re = re.compile(r'^(\d+)\.\s')

    for i, child in enumerate(children):
        if child.tag == 'P':
            text = get_text(child).strip()
            m = note_number_re.match(text)
            if m:
                note_num = int(m.group(1))
                note_starts.append((i, note_num))

    if not note_starts:
        return [(1, children)] if children else []

    # 각 주석의 요소 범위 결정
    raw_notes = []  # (번호, [요소들])
    for idx, (start_i, note_num) in enumerate(note_starts):
        if idx + 1 < len(note_starts):
            end_i = note_starts[idx + 1][0]
        else:
            end_i = len(children)

        note_elements = []
        for i in range(start_i, end_i):
            child = children[i]
            if child.tag == 'PGBRK':
                continue
            note_elements.append(child)

        if note_elements:
            raw_notes.append((note_num, note_elements))

    # max_sheet_num 이후의 주석을 마지막 시트에 합침
    notes = []
    for note_num, elems in raw_notes:
        if note_num <= max_sheet_num:
            notes.append((note_num, elems))
        else:
            # 마지막 시트에 합침
            if notes:
                notes[-1][1].extend(elems)
            else:
                notes.append((note_num, elems))

    return notes


# ============================================================================
# 5단계: Excel 시트 생성
# ============================================================================

def write_cover_sheet(wb, cover_info):
    """Cover 시트 생성"""
    ws = wb.create_sheet('Cover')
    font_title = Font(name=DEFAULT_FONT_NAME, size=18)
    font_normal = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    center = Alignment(horizontal='center', vertical='center')

    data = [
        (1, cover_info['company_name'], font_title),
        (3, cover_info['report_subtitle'], font_title),
        (5, cover_info['report_title'], font_title),
        (6, cover_info['period_line1'], font_normal),
        (7, (cover_info['period_from'] + ' 부터') if cover_info['period_from'] else '', font_normal),
        (8, (cover_info['period_to'] + ' 까지') if cover_info['period_to'] else '', font_normal),
        (10, cover_info['auditor_name'], font_title),
    ]

    for row, text, font in data:
        if text:
            cell = ws.cell(row=row, column=DATA_COL_START, value=text)
            cell.font = font
            cell.alignment = center

    if cover_info['period_line1']:
        ws.merge_cells(start_row=6, start_column=DATA_COL_START,
                       end_row=6, end_column=DATA_COL_START + 1)

    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    return ws


def write_toc_sheet(wb, toc_items):
    """ToC 시트 생성"""
    ws = wb.create_sheet('ToC')
    font_title = Font(name=DEFAULT_FONT_NAME, size=12, bold=True)
    font_normal = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)

    row = 1
    for item in toc_items:
        if item[0] == 'title':
            cell = ws.cell(row=row, column=DATA_COL_START, value=item[1])
            cell.font = font_title
            cell.alignment = Alignment(horizontal='center')
        elif item[0] == 'entry':
            text = item[1]
            if len(item) > 2 and item[2]:
                text = f"{item[1]}  {item[2]}"
            cell = ws.cell(row=row, column=DATA_COL_START, value=text)
            cell.font = font_normal
        row += 1

    ws.column_dimensions['D'].width = 60
    return ws


def write_opinion_sheet(wb, opinion_section):
    """Opinion 시트 생성"""
    ws = wb.create_sheet('Opinion')
    if opinion_section is None:
        return ws

    font_normal = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    font_bold = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    row = 1

    for child in opinion_section:
        tag = child.tag
        if tag == 'TITLE':
            cell = ws.cell(row=row, column=DATA_COL_START, value=get_text(child).strip())
            cell.font = Font(name=DEFAULT_FONT_NAME, size=12, bold=True)
            row += 1
            continue

        if tag == 'P':
            text = get_text(child).strip()
            if not text:
                row += 1
                continue

            usermark = child.get('USERMARK', '')
            info = parse_usermark(usermark)
            is_bold = info.get('bold', False)

            lines = text.split('\n')
            for line in lines:
                line = line.strip()
                if line:
                    cell = ws.cell(row=row, column=DATA_COL_START, value=line)
                    cell.font = font_bold if is_bold else font_normal
                    cell.alignment = Alignment(wrap_text=True)
                row += 1
            continue

        if tag == 'TABLE':
            row = write_table_to_sheet(ws, child, row, DATA_COL_START)

    ws.column_dimensions['D'].width = 80
    return ws


def write_fs_sheet(wb, header_elements):
    """FS(재무제표 표지) 시트 생성"""
    ws = wb.create_sheet('FS')
    row = 1

    for elem in header_elements:
        tag = elem.tag
        if tag == 'P':
            text = get_text(elem).strip()
            if text:
                usermark = elem.get('USERMARK', '')
                info = parse_usermark(usermark)
                size = info.get('size', DEFAULT_FONT_SIZE)
                bold = info.get('bold', False)
                cell = ws.cell(row=row, column=DATA_COL_START, value=text)
                cell.font = Font(name=DEFAULT_FONT_NAME, size=size, bold=bold)
            row += 1
        elif tag == 'TABLE':
            row = write_table_to_sheet(ws, elem, row, DATA_COL_START)
        elif tag == 'TABLE-GROUP':
            for t in elem.findall('TABLE'):
                row = write_table_to_sheet(ws, t, row, DATA_COL_START)

    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    return ws


def write_financial_statement_sheet(wb, sheet_name, elements, col_start=DATA_COL_START):
    """재무제표 시트 생성"""
    ws = wb.create_sheet(sheet_name)
    row = 1

    for elem in elements:
        tag = elem.tag
        if tag == 'P':
            text = get_text(elem).strip()
            if text:
                usermark = elem.get('USERMARK', '')
                info = parse_usermark(usermark)
                size = info.get('size', DEFAULT_FONT_SIZE)
                bold = info.get('bold', False)
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    if line:
                        cell = ws.cell(row=row, column=col_start, value=line)
                        cell.font = Font(name=DEFAULT_FONT_NAME, size=size, bold=bold)
                        cell.alignment = Alignment(horizontal='center')
                    row += 1
            else:
                row += 1

        elif tag == 'TABLE':
            table_data = parse_table(elem)
            row = write_parsed_table(ws, table_data, row, col_start)

        elif tag == 'TABLE-GROUP':
            for t in elem.findall('TABLE'):
                table_data = parse_table(t)
                row = write_parsed_table(ws, table_data, row, col_start)

    # 열 너비 설정
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 31
    ws.column_dimensions['E'].width = 21
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 5

    return ws


def write_financial_statement_divided(wb, orig_sheet_name, div_sheet_name):
    """천원 단위 재무제표 시트 생성"""
    orig_ws = wb[orig_sheet_name]
    ws = wb.create_sheet(div_sheet_name)

    for row in orig_ws.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column)

            if isinstance(cell.value, (int, float)) and cell.value != 0:
                divided = round(cell.value / FS_DIVIDER)
                new_cell.value = divided if divided != 0 else cell.value
            else:
                new_cell.value = cell.value

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.number_format = cell.number_format

    for merge_range in orig_ws.merged_cells.ranges:
        ws.merge_cells(str(merge_range))

    for col_letter, dim in orig_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
        ws.column_dimensions[col_letter].hidden = dim.hidden

    # 단위 텍스트 수정
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '단위' in cell.value and '원' in cell.value:
                cell.value = cell.value.replace('단위 : 원', '단위 : 천원').replace('단위: 원', '단위: 천원')

    return ws


def write_fn_sheet(wb, cover_info):
    """FN(주석 표지) 시트 생성"""
    ws = wb.create_sheet('FN')
    font_title = Font(name=DEFAULT_FONT_NAME, size=12, bold=True)
    font_normal = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)

    ws.cell(row=1, column=DATA_COL_START, value='주석').font = font_title

    if cover_info.get('period_from') and cover_info.get('period_to'):
        ws.cell(row=2, column=DATA_COL_START,
                value=f"제46(당)기 {cover_info['period_from']} 부터 {cover_info['period_to']} 까지").font = font_normal
        ws.cell(row=3, column=DATA_COL_START,
                value="제45(전)기").font = font_normal

    ws.cell(row=5, column=DATA_COL_START,
            value=cover_info.get('company_name', '')).font = font_normal

    ws.column_dimensions['D'].width = 60
    return ws


def write_note_sheet(wb, sheet_name, note_elements):
    """개별 주석 시트 생성"""
    ws = wb.create_sheet(sheet_name)
    font_normal = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    row = 1

    for elem in note_elements:
        tag = elem.tag
        if tag == 'P':
            text = get_text(elem).strip()
            if not text:
                row += 1
                continue

            usermark = elem.get('USERMARK', '')
            info = parse_usermark(usermark)
            size = info.get('size', DEFAULT_FONT_SIZE)
            bold = info.get('bold', False)

            lines = text.split('\n')
            for line in lines:
                line = line.strip()
                if line:
                    cell = ws.cell(row=row, column=DATA_COL_START, value=line)
                    cell.font = Font(name=DEFAULT_FONT_NAME, size=size, bold=bold)
                    cell.alignment = Alignment(wrap_text=True)
                row += 1

        elif tag == 'TABLE':
            table_data = parse_table(elem)
            row = write_parsed_table(ws, table_data, row, DATA_COL_START)

        elif tag == 'TABLE-GROUP':
            for t in elem.findall('TABLE'):
                table_data = parse_table(t)
                row = write_parsed_table(ws, table_data, row, DATA_COL_START)

    ws.column_dimensions['D'].width = 35
    for col_idx in range(5, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    return ws


def write_conduct_sheet(wb, conduct_section):
    """Conduct 시트 생성"""
    ws = wb.create_sheet('Conduct')
    if conduct_section is None:
        return ws

    font_normal = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    row = 1

    for child in conduct_section:
        tag = child.tag
        if tag == 'TITLE':
            cell = ws.cell(row=row, column=DATA_COL_START, value=get_text(child).strip())
            cell.font = Font(name=DEFAULT_FONT_NAME, size=12, bold=True)
            row += 1

        elif tag == 'SECTION-2':
            for sec_child in child:
                sec_tag = sec_child.tag
                if sec_tag == 'TITLE':
                    cell = ws.cell(row=row, column=DATA_COL_START, value=get_text(sec_child).strip())
                    cell.font = Font(name=DEFAULT_FONT_NAME, size=10, bold=True)
                    row += 1
                elif sec_tag == 'TABLE':
                    table_data = parse_table(sec_child)
                    row = write_parsed_table(ws, table_data, row, DATA_COL_START)
                elif sec_tag == 'TABLE-GROUP':
                    for t in sec_child.findall('TABLE'):
                        table_data = parse_table(t)
                        row = write_parsed_table(ws, table_data, row, DATA_COL_START)
                elif sec_tag == 'P':
                    text = get_text(sec_child).strip()
                    if text:
                        lines = text.split('\n')
                        for line in lines:
                            line = line.strip()
                            if line:
                                ws.cell(row=row, column=DATA_COL_START, value=line).font = font_normal
                            row += 1
                    else:
                        row += 1

        elif tag == 'P':
            text = get_text(child).strip()
            if text:
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    if line:
                        ws.cell(row=row, column=DATA_COL_START, value=line).font = font_normal
                    row += 1
            else:
                row += 1

        elif tag == 'TABLE':
            table_data = parse_table(child)
            row = write_parsed_table(ws, table_data, row, DATA_COL_START)

        elif tag == 'TABLE-GROUP':
            for t in child.findall('TABLE'):
                table_data = parse_table(t)
                row = write_parsed_table(ws, table_data, row, DATA_COL_START)

    ws.column_dimensions['D'].width = 20
    for col_idx in range(5, 20):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    return ws


# ============================================================================
# 테이블을 시트에 기록하는 공통 함수
# ============================================================================

def write_table_to_sheet(ws, tag_elem, start_row, start_col):
    """TABLE XML 요소를 직접 파싱하여 시트에 기록"""
    table_data = parse_table(tag_elem)
    return write_parsed_table(ws, table_data, start_row, start_col)


def write_parsed_table(ws, table_data, start_row, start_col):
    """파싱된 테이블 데이터를 시트에 기록하고 다음 행 번호 반환"""
    has_border = table_data['border']
    col_widths = table_data['col_widths']
    all_rows = table_data['thead_rows'] + table_data['tbody_rows']
    num_thead = len(table_data['thead_rows'])

    if not all_rows:
        return start_row

    # 열 너비 설정
    for i, w in enumerate(col_widths):
        col_letter = get_column_letter(start_col + i)
        current = ws.column_dimensions[col_letter].width or 0
        new_width = px_to_excel_width(w)
        if new_width > current:
            ws.column_dimensions[col_letter].width = new_width

    # 셀 점유 맵 (rowspan/colspan 처리)
    occupied = {}

    row_offset = 0
    for row_idx, row_cells in enumerate(all_rows):
        is_thead_row = row_idx < num_thead
        col_offset = 0

        for cell_data in row_cells:
            while (row_offset, col_offset) in occupied:
                col_offset += 1

            colspan = cell_data['colspan']
            rowspan = cell_data['rowspan']
            text = cell_data['text']
            is_header = cell_data['is_header'] or is_thead_row

            value, is_number = try_parse_number(text)

            excel_row = start_row + row_offset
            excel_col = start_col + col_offset
            cell = ws.cell(row=excel_row, column=excel_col, value=value)

            # 폰트
            usermark = cell_data.get('usermark', '')
            um_info = parse_usermark(usermark)
            font_size = um_info.get('size', DEFAULT_FONT_SIZE)
            font_bold = um_info.get('bold', is_header)
            cell.font = Font(name=DEFAULT_FONT_NAME, size=font_size, bold=font_bold)

            # 정렬
            h_align = cell_data.get('align', '').lower() or None
            v_align = cell_data.get('valign', '').lower() or None
            if h_align == 'justify':
                h_align = 'left'
            if v_align == 'middle':
                v_align = 'center'
            if is_header and not h_align:
                h_align = 'center'
            cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

            # 숫자 포맷
            if is_number:
                cell.number_format = NUMBER_FORMAT

            # 테두리
            if has_border:
                cell.border = THIN_BORDER

            # 병합
            if colspan > 1 or rowspan > 1:
                end_row = excel_row + rowspan - 1
                end_col = excel_col + colspan - 1
                ws.merge_cells(
                    start_row=excel_row, start_column=excel_col,
                    end_row=end_row, end_column=end_col
                )
                if has_border:
                    for r in range(excel_row, end_row + 1):
                        for c in range(excel_col, end_col + 1):
                            ws.cell(row=r, column=c).border = THIN_BORDER

                for dr in range(rowspan):
                    for dc in range(colspan):
                        if dr == 0 and dc == 0:
                            continue
                        occupied[(row_offset + dr, col_offset + dc)] = True

            col_offset += colspan

        row_offset += 1

    return start_row + row_offset


# ============================================================================
# 메인
# ============================================================================

def convert_dsd_to_excel(dsd_path, output_path=None):
    """DSD 파일을 Excel로 변환"""
    if output_path is None:
        base = os.path.splitext(dsd_path)[0]
        output_path = base + '.xlsx'

    print(f"DSD 파일 읽는 중: {dsd_path}")
    meta_root, contents_root = parse_dsd(dsd_path)

    print("문서 구조 분석 중...")
    doc = analyze_document(contents_root)

    cover_info = extract_cover_info(doc['cover'])
    print(f"  회사명: {cover_info['company_name']}")
    print(f"  보고서: {cover_info['report_title']}")
    print(f"  감사법인: {cover_info['auditor_name']}")

    toc_items = extract_toc_items(doc['toc'])

    statements = split_financial_statements(doc['fs_section'])
    print(f"  재무상태표 요소: {len(statements['bs'])}개")
    print(f"  포괄손익계산서 요소: {len(statements['is'])}개")
    print(f"  자본변동표 요소: {len(statements['ce'])}개")
    print(f"  현금흐름표 요소: {len(statements['cf'])}개")

    notes = split_notes(doc['notes_section'])
    print(f"  주석: {len(notes)}개")

    # Excel 워크북 생성
    print("\nExcel 워크북 생성 중...")
    wb = Workbook()
    wb.remove(wb.active)

    print("  Cover 시트...")
    write_cover_sheet(wb, cover_info)

    print("  ToC 시트...")
    write_toc_sheet(wb, toc_items)

    print("  Opinion 시트...")
    write_opinion_sheet(wb, doc['opinion_section'])

    print("  FS 시트...")
    write_fs_sheet(wb, statements['fs_header'])

    # 재무제표
    fs_pairs = [
        ('BS', 'BS2', statements['bs']),
        ('IS', 'IS2', statements['is']),
        ('CF', 'CF2', statements['cf']),
        ('CE', 'CE2', statements['ce']),
    ]
    for name, name2, elems in fs_pairs:
        print(f"  {name} 시트...")
        write_financial_statement_sheet(wb, name, elems)
        print(f"  {name2} 시트...")
        write_financial_statement_divided(wb, name, name2)

    print("  FN 시트...")
    write_fn_sheet(wb, cover_info)

    # 주석
    for note_num, note_elems in notes:
        print(f"  주석 {note_num} 시트...")
        write_note_sheet(wb, str(note_num), note_elems)

    # Sox 빈 시트
    wb.create_sheet('Sox')

    print("  Conduct 시트...")
    write_conduct_sheet(wb, doc['conduct_section'])

    print(f"\n저장 중: {output_path}")
    wb.save(output_path)
    print(f"완료! {output_path}")
    print(f"총 시트 수: {len(wb.sheetnames)}")
    print(f"시트: {', '.join(wb.sheetnames)}")


def main():
    if len(sys.argv) < 2:
        print("사용법: python dsd_to_excel.py <input.dsd> [output.xlsx]")
        sys.exit(1)

    dsd_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(dsd_path):
        print(f"오류: 파일을 찾을 수 없습니다: {dsd_path}")
        sys.exit(1)

    convert_dsd_to_excel(dsd_path, output_path)


if __name__ == '__main__':
    main()
